from __future__ import annotations
'\nVCdb data export module.\n\nThis module provides functionality for exporting query results to different file formats,\nincluding CSV and Excel.\n'
import csv
import logging
import os
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple, Union, Awaitable
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
class ExportError(Exception):
    def __init__(self, message: str, details: Optional[Dict[str, Any]]=None) -> None:
        self.message = message
        self.details = details or {}
        super().__init__(message)
class DataExporter:
    def __init__(self, logger: logging.Logger) -> None:
        self._logger = logger
    async def export_csv(self, data: List[Dict[str, Any]], columns: List[str], column_map: Dict[str, str], file_path: str) -> None:
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=columns, extrasaction='ignore')
                writer.writerow({col: column_map.get(col, col) for col in columns})
                for row in data:
                    clean_row = {}
                    for col in columns:
                        value = row.get(col, '')
                        clean_row[col] = str(value) if value is not None else ''
                    writer.writerow(clean_row)
            self._logger.info(f'Exported {len(data)} rows to CSV: {file_path}')
        except Exception as e:
            self._logger.error(f'Failed to export CSV: {str(e)}')
            raise ExportError(f'Failed to export CSV: {str(e)}') from e
    async def export_excel(self, data: List[Dict[str, Any]], columns: List[str], column_map: Dict[str, str], file_path: str, sheet_name: Optional[str]=None) -> None:
        if not EXCEL_AVAILABLE:
            raise ExportError('Excel export is not available. Please install openpyxl.')
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if not sheet_name:
                sheet_name = 'VCdb Results'
            ws.title = sheet_name
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=1, column=1, value=f'Generated: {timestamp}')
            ws.cell(row=1, column=1).font = Font(italic=True)
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            header_alignment = Alignment(horizontal='center')
            for col_idx, col_id in enumerate(columns, 1):
                cell = ws.cell(row=2, column=col_idx, value=column_map.get(col_id, col_id))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            for row_idx, row_data in enumerate(data, 3):
                for col_idx, col_id in enumerate(columns, 1):
                    value = row_data.get(col_id, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            for col_idx, _ in enumerate(columns, 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(2, min(len(data) + 3, 100)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max(max_length + 2, 10)
                ws.column_dimensions[col_letter].width = min(adjusted_width, 50)
            ws.freeze_panes = 'A3'
            wb.save(file_path)
            self._logger.info(f'Exported {len(data)} rows to Excel: {file_path}')
        except Exception as e:
            self._logger.error(f'Failed to export Excel: {str(e)}')
            raise ExportError(f'Failed to export Excel: {str(e)}') from e
    async def export_all_data(self, database_callback: Callable[..., Awaitable[Tuple[List[Dict[str, Any]], int]]], filter_panels: List[Dict[str, List[int]]], columns: List[str], column_map: Dict[str, str], file_path: str, format_type: str, max_rows: int=0, table_filters: Optional[Dict[str, Any]]=None, sort_by: Optional[str]=None, sort_desc: bool=False, progress_callback: Optional[Callable[[int, int], Awaitable[Optional[bool]]]]=None) -> int:
        try:
            results, total_count = await database_callback(filter_panels=filter_panels, columns=columns, page=1, page_size=1, sort_by=sort_by, sort_desc=sort_desc, table_filters=table_filters)
            if max_rows == 0:
                max_rows = total_count
            total_to_export = min(total_count, max_rows)
            if total_to_export == 0:
                self._logger.warning('No data to export')
                return 0
            if format_type == 'csv':
                return await self._export_all_csv(database_callback, filter_panels, columns, column_map, file_path, total_to_export, table_filters, sort_by, sort_desc, progress_callback)
            elif format_type == 'excel' and EXCEL_AVAILABLE:
                return await self._export_all_excel(database_callback, filter_panels, columns, column_map, file_path, total_to_export, table_filters, sort_by, sort_desc, progress_callback)
            else:
                raise ExportError(f'Unsupported export format: {format_type}')
        except Exception as e:
            self._logger.error(f'Failed to export data: {str(e)}')
            raise ExportError(f'Failed to export data: {str(e)}') from e
    async def _export_all_csv(self, database_callback: Callable[..., Awaitable[Tuple[List[Dict[str, Any]], int]]], filter_panels: List[Dict[str, List[int]]], columns: List[str], column_map: Dict[str, str], file_path: str, total_to_export: int, table_filters: Optional[Dict[str, Any]], sort_by: Optional[str], sort_desc: bool, progress_callback: Optional[Callable[[int, int], Awaitable[Optional[bool]]]]) -> int:
        with open(file_path, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=columns, extrasaction='ignore')
            writer.writerow({col: column_map.get(col, col) for col in columns})
            batch_size = 1000
            rows_exported = 0
            for page in range(1, (total_to_export + batch_size - 1) // batch_size + 1):
                if progress_callback:
                    cancel = await progress_callback(-1, total_to_export)
                    if cancel is False:
                        return rows_exported
                results, _ = await database_callback(filter_panels=filter_panels, columns=columns, page=page, page_size=batch_size, sort_by=sort_by, sort_desc=sort_desc, table_filters=table_filters)
                for row in results:
                    clean_row = {}
                    for col in columns:
                        value = row.get(col, '')
                        clean_row[col] = str(value) if value is not None else ''
                    writer.writerow(clean_row)
                    rows_exported += 1
                    if progress_callback:
                        cancel = await progress_callback(rows_exported, total_to_export)
                        if cancel is False:
                            return rows_exported
                    if rows_exported >= total_to_export:
                        break
        self._logger.info(f'Exported {rows_exported} rows to CSV: {file_path}')
        return rows_exported
    async def _export_all_excel(self, database_callback: Callable[..., Awaitable[Tuple[List[Dict[str, Any]], int]]], filter_panels: List[Dict[str, List[int]]], columns: List[str], column_map: Dict[str, str], file_path: str, total_to_export: int, table_filters: Optional[Dict[str, Any]], sort_by: Optional[str], sort_desc: bool, progress_callback: Optional[Callable[[int, int], Awaitable[Optional[bool]]]]) -> int:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'VCdb Results'
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=1, column=1, value=f'Generated: {timestamp}')
        ws.cell(row=1, column=1).font = Font(italic=True)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        header_alignment = Alignment(horizontal='center')
        for col_idx, col_id in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=column_map.get(col_id, col_id))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        batch_size = 1000
        rows_exported = 0
        excel_row = 3
        for page in range(1, (total_to_export + batch_size - 1) // batch_size + 1):
            if progress_callback:
                cancel = await progress_callback(-1, total_to_export)
                if cancel is False:
                    return rows_exported
            results, _ = await database_callback(filter_panels=filter_panels, columns=columns, page=page, page_size=batch_size, sort_by=sort_by, sort_desc=sort_desc, table_filters=table_filters)
            for row_data in results:
                for col_idx, col_id in enumerate(columns, 1):
                    value = row_data.get(col_id, '')
                    ws.cell(row=excel_row, column=col_idx, value=value)
                excel_row += 1
                rows_exported += 1
                if progress_callback:
                    cancel = await progress_callback(rows_exported, total_to_export)
                    if cancel is False:
                        return rows_exported
                if rows_exported >= total_to_export:
                    break
        for col_idx, _ in enumerate(columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(2, min(excel_row, 102)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 10)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)
        ws.freeze_panes = 'A3'
        wb.save(file_path)
        self._logger.info(f'Exported {rows_exported} rows to Excel: {file_path}')
        return rows_exported