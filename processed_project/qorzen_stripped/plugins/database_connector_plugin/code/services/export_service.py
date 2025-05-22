from __future__ import annotations
import csv
import io
import json
import logging
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Any, Dict, List, Optional
from ..models import QueryResult, ExportSettings, ExportFormat
class ExportService:
    def __init__(self, file_manager: Any, logger: logging.Logger) -> None:
        self._file_manager = file_manager
        self._logger = logger
    async def export_results(self, results: QueryResult, format: ExportFormat, file_path: str, settings: ExportSettings) -> str:
        try:
            self._logger.info(f'Exporting {results.row_count} rows to {format.value} format')
            data = self._prepare_data(results, settings)
            if format == ExportFormat.CSV:
                content = self._export_to_csv(data, settings)
                await self._file_manager.write_text(file_path, content, create_dirs=True)
            elif format == ExportFormat.JSON:
                content = self._export_to_json(data, settings)
                await self._file_manager.write_text(file_path, content, create_dirs=True)
            elif format == ExportFormat.XML:
                content = self._export_to_xml(data, settings)
                await self._file_manager.write_text(file_path, content, create_dirs=True)
            elif format == ExportFormat.TSV:
                content = self._export_to_tsv(data, settings)
                await self._file_manager.write_text(file_path, content, create_dirs=True)
            elif format == ExportFormat.HTML:
                content = self._export_to_html(data, results, settings)
                await self._file_manager.write_text(file_path, content, create_dirs=True)
            elif format == ExportFormat.EXCEL:
                content = self._export_to_excel(data, settings)
                await self._file_manager.write_binary(file_path, content, create_dirs=True)
            else:
                raise ValueError(f'Unsupported export format: {format}')
            self._logger.info(f'Successfully exported to {file_path}')
            return file_path
        except Exception as e:
            self._logger.error(f'Failed to export results: {e}')
            raise
    def _prepare_data(self, results: QueryResult, settings: ExportSettings) -> List[Dict[str, Any]]:
        data = results.records.copy()
        if settings.max_rows and len(data) > settings.max_rows:
            data = data[:settings.max_rows]
            self._logger.info(f'Limited export to {settings.max_rows} rows')
        for row in data:
            for key, value in row.items():
                if value is None:
                    row[key] = settings.null_value
                elif isinstance(value, datetime):
                    row[key] = value.strftime(settings.date_format)
                elif isinstance(value, (int, float)) and str(value).lower() in ('nan', 'inf', '-inf'):
                    row[key] = settings.null_value
        return data
    def _export_to_csv(self, data: List[Dict[str, Any]], settings: ExportSettings) -> str:
        if not data:
            return ''
        output = io.StringIO()
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=settings.delimiter, quotechar=settings.quote_char, quoting=csv.QUOTE_MINIMAL)
        if settings.include_headers:
            writer.writeheader()
        writer.writerows(data)
        return output.getvalue()
    def _export_to_tsv(self, data: List[Dict[str, Any]], settings: ExportSettings) -> str:
        if not data:
            return ''
        output = io.StringIO()
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter='\t', quotechar=settings.quote_char, quoting=csv.QUOTE_MINIMAL)
        if settings.include_headers:
            writer.writeheader()
        writer.writerows(data)
        return output.getvalue()
    def _export_to_json(self, data: List[Dict[str, Any]], settings: ExportSettings) -> str:
        return json.dumps(data, indent=2, ensure_ascii=False, default=str)
    def _export_to_xml(self, data: List[Dict[str, Any]], settings: ExportSettings) -> str:
        root = ET.Element('data')
        for row in data:
            record_element = ET.SubElement(root, 'record')
            for key, value in row.items():
                field_element = ET.SubElement(record_element, 'field')
                field_element.set('name', str(key))
                field_element.text = str(value) if value is not None else settings.null_value
        self._indent_xml(root)
        return '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(root, encoding='unicode')
    def _export_to_html(self, data: List[Dict[str, Any]], results: QueryResult, settings: ExportSettings) -> str:
        if not data:
            return '<html><body><p>No data to display</p></body></html>'
        html_parts = ['<!DOCTYPE html>', '<html>', '<head>', "    <meta charset='UTF-8'>", '    <title>Query Results Export</title>', '    <style>', '        body { font-family: Arial, sans-serif; margin: 20px; }', '        table { border-collapse: collapse; width: 100%; }', '        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }', '        th { background-color: #f2f2f2; font-weight: bold; }', '        tr:nth-child(even) { background-color: #f9f9f9; }', '        .metadata { background-color: #e9f4ff; padding: 10px; margin-bottom: 20px; border-radius: 5px; }', '        .metadata h3 { margin-top: 0; }', '    </style>', '</head>', '<body>', "    <div class='metadata'>", '        <h3>Query Results</h3>', f'        <p><strong>Executed:</strong> {results.executed_at.strftime(settings.date_format)}</p>', f'        <p><strong>Connection:</strong> {results.connection_id}</p>', f'        <p><strong>Rows:</strong> {results.row_count}</p>', f'        <p><strong>Execution Time:</strong> {results.execution_time_ms}ms</p>']
        if results.truncated:
            html_parts.append('        <p><strong>Note:</strong> Results may be truncated</p>')
        html_parts.extend(['    </div>', '    <table>'])
        if settings.include_headers:
            html_parts.append('        <thead>')
            html_parts.append('            <tr>')
            for key in data[0].keys():
                html_parts.append(f'                <th>{self._html_escape(str(key))}</th>')
            html_parts.append('            </tr>')
            html_parts.append('        </thead>')
        html_parts.append('        <tbody>')
        for row in data:
            html_parts.append('            <tr>')
            for value in row.values():
                escaped_value = self._html_escape(str(value) if value is not None else settings.null_value)
                html_parts.append(f'                <td>{escaped_value}</td>')
            html_parts.append('            </tr>')
        html_parts.append('        </tbody>')
        html_parts.extend(['    </table>', '</body>', '</html>'])
        return '\n'.join(html_parts)
    def _export_to_excel(self, data: List[Dict[str, Any]], settings: ExportSettings) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self._logger.warning('openpyxl not available, creating simple Excel-compatible CSV')
            csv_content = self._export_to_csv(data, settings)
            return csv_content.encode(settings.encoding)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Query Results'
        if not data:
            worksheet['A1'] = 'No data to export'
            output = io.BytesIO()
            workbook.save(output)
            return output.getvalue()
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='E9F4FF', end_color='E9F4FF', fill_type='solid')
        if settings.include_headers:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
            start_row = 2
        else:
            start_row = 1
        for row_idx, row in enumerate(data, start_row):
            for col_idx, (key, value) in enumerate(row.items(), 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if value is None:
                    cell.value = settings.null_value
                elif isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value)
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()
    def _indent_xml(self, element: ET.Element, level: int=0) -> None:
        indent = '\n' + level * '  '
        if len(element):
            if not element.text or not element.text.strip():
                element.text = indent + '  '
            if not element.tail or not element.tail.strip():
                element.tail = indent
            for child in element:
                self._indent_xml(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = indent
        elif level and (not element.tail or not element.tail.strip()):
            element.tail = indent
    def _html_escape(self, text: str) -> str:
        escape_dict = {'&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#x27;'}
        for char, escape in escape_dict.items():
            text = text.replace(char, escape)
        return text
    def get_supported_formats(self) -> List[ExportFormat]:
        return list(ExportFormat)
    def get_file_extension(self, format: ExportFormat) -> str:
        extensions = {ExportFormat.CSV: '.csv', ExportFormat.JSON: '.json', ExportFormat.XML: '.xml', ExportFormat.EXCEL: '.xlsx', ExportFormat.TSV: '.tsv', ExportFormat.HTML: '.html'}
        return extensions.get(format, '.txt')
    def get_mime_type(self, format: ExportFormat) -> str:
        mime_types = {ExportFormat.CSV: 'text/csv', ExportFormat.JSON: 'application/json', ExportFormat.XML: 'application/xml', ExportFormat.EXCEL: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', ExportFormat.TSV: 'text/tab-separated-values', ExportFormat.HTML: 'text/html'}
        return mime_types.get(format, 'text/plain')