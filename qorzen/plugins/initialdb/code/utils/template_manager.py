from __future__ import annotations

"""
Template management module for InitialDB application.

This module handles loading, saving, and exporting data to various templates,
including ACES template format for automotive data interchange.
"""

import os
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union, TypedDict, cast
import csv
from enum import Enum, auto
import structlog
from openpyxl import Workbook

from ..config.settings import settings, TEMPLATES_DIR

logger = structlog.get_logger(__name__)


class TemplateType(Enum):
    """Enum defining supported template formats."""
    ACES = auto()
    CUSTOM = auto()


class FieldMapping(TypedDict):
    """TypedDict defining the structure of field mapping entries."""
    model: Optional[str]
    attribute: Optional[str]


class TemplateManager:
    """
    Manager class for handling data templates.

    Provides functionality for loading template definitions, mapping data to templates,
    and exporting data in template formats.
    """

    def __init__(self) -> None:
        """Initialize the template manager with available templates."""
        self.templates_dir = TEMPLATES_DIR
        self.templates: Dict[str, Dict[str, FieldMapping]] = {}
        self._load_templates()

    def _load_templates(self) -> None:
        """
        Load all template definitions from the templates directory.

        Each template is a JSON file containing field mappings.
        """
        if not os.path.exists(self.templates_dir):
            os.makedirs(self.templates_dir, exist_ok=True)

        # Create default ACES template if none exists
        aces_template_path = os.path.join(self.templates_dir, "ACES_Template.json")
        if not os.path.exists(aces_template_path):
            self._create_default_aces_template(aces_template_path)

        # Load all template files
        template_files = os.listdir(self.templates_dir)

        # If no templates exist at all, create the default ACES template
        if not any(file.endswith('.json') for file in template_files):
            self._create_default_aces_template(aces_template_path)

        # Load all template files
        for file in os.listdir(self.templates_dir):
            if file.endswith('.json'):
                template_path = os.path.join(self.templates_dir, file)
                template_name = os.path.splitext(file)[0]
                try:
                    with open(template_path, 'r', encoding='utf-8') as f:
                        template_data = json.load(f)
                        self.templates[template_name] = template_data
                        logger.info(f"Loaded template: {template_name}")
                except Exception as e:
                    logger.error(f"Error loading template {template_name}: {str(e)}")

        # If still no templates loaded, create and load the default ACES template
        if not self.templates:
            self._create_default_aces_template(aces_template_path)
            try:
                with open(aces_template_path, 'r', encoding='utf-8') as f:
                    template_data = json.load(f)
                    self.templates["ACES_Template"] = template_data
                    logger.info("Loaded default ACES template")
            except Exception as e:
                logger.error(f"Error loading default ACES template: {str(e)}")

    def _create_default_aces_template(self, path: str) -> None:
        """
        Create the default ACES template file if it doesn't exist.

        Args:
            path: Path where the template file should be created
        """
        default_aces_template = {
            "Make": {"model": "Make", "attribute": "make_name"},
            "Model": {"model": "Model", "attribute": "model_name"},
            "Year": {"model": "Year", "attribute": "year_id"},
            "VehicleType": {"model": "VehicleType", "attribute": "vehicle_type_name"},
            "SubModel": {"model": "SubModel", "attribute": "sub_model_name"},
            "Product": {"model": None, "attribute": None},
            "PartNumber": {"model": None, "attribute": None},
            "MfrLabel": {"model": None, "attribute": None},
            "Position": {"model": None, "attribute": None},
            "Notes": {"model": None, "attribute": None},
            "EngineLiters": {"model": "EngineBlock", "attribute": "liter"},
            "EngineCC": {"model": "EngineBlock", "attribute": "cc"},
            "EngCoreCID": {"model": "EngineBlock", "attribute": "cid"},
            "EngineCylinder": {"model": "EngineBlock", "attribute": "cylinders"},
            "EngineBlock": {"model": "EngineBlock", "attribute": "block_type"},
            "EngBoreInch": {"model": "EngineBoreStroke", "attribute": "eng_bore_in"},
            "EngBoreMetric": {"model": "EngineBoreStroke", "attribute": "eng_bore_metric"},
            "EngStrokeInch": {"model": "EngineBoreStroke", "attribute": "eng_stroke_in"},
            "EngStrokeMetric": {"model": "EngineBoreStroke", "attribute": "eng_stroke_metric"},
            "Qty": {"model": None, "attribute": None},
            "BodyNumDoors": {"model": "BodyNumDoors", "attribute": "body_num_doors"},
            "BodyType": {"model": "BodyType", "attribute": "body_type_name"},
            "BedLengthIn": {"model": "BedLength", "attribute": "bed_length"},
            "BedLengthMetric": {"model": "BedLength", "attribute": "bed_length_metric"},
            "BedType": {"model": "BedType", "attribute": "bed_type_name"},
            "DriveType": {"model": "DriveType", "attribute": "drive_type_name"},
            "MfrBodyCode": {"model": "MfrBodyCode", "attribute": "mfr_body_code_name"},
            "WheelBaseIn": {"model": "WheelBase", "attribute": "wheel_base"},
            "WheelBaseMetric": {"model": "WheelBase", "attribute": "wheel_base_metric"},
            "FrontBrakeType": {"model": "BrakeConfig", "attribute": "front_brake_type.brake_type_name"},
            "RearBrakeType": {"model": "BrakeConfig", "attribute": "rear_brake_type.brake_type_name"},
            "BrakeSystem": {"model": "BrakeSystem", "attribute": "brake_system_name"},
            "BrakeABS": {"model": "BrakeABS", "attribute": "brake_abs_name"},
            "FrontSpringType": {"model": "SpringTypeConfig", "attribute": "front_spring_type.spring_type_name"},
            "RearSpringType": {"model": "SpringTypeConfig", "attribute": "rear_spring_type.spring_type_name"},
            "SteeringType": {"model": "SteeringType", "attribute": "steering_type_name"},
            "SteeringSystem": {"model": "SteeringSystem", "attribute": "steering_system_name"},
            "EngineDesignation": {"model": "EngineDesignation", "attribute": "engine_designation_name"},
            "EngineVIN": {"model": "EngineVIN", "attribute": "engine_vin_name"},
            "Aspiration": {"model": "Aspiration", "attribute": "aspiration_name"},
            "CylinderHeadType": {"model": "CylinderHeadType", "attribute": "cylinder_head_type_name"},
            "FuelType": {"model": "FuelType", "attribute": "fuel_type_name"},
            "IgnitionSystemType": {"model": "IgnitionSystemType", "attribute": "ignition_system_type_name"},
            "EngineVersion": {"model": "EngineVersion", "attribute": "engine_version"},
            "FuelDeliveryType": {"model": "FuelDeliveryType", "attribute": "fuel_delivery_type_name"},
            "FuelDeliverySubType": {"model": "FuelDeliverySubType", "attribute": "fuel_delivery_sub_type_name"},
            "FuelSystemControlType": {"model": "FuelSystemControlType", "attribute": "fuel_system_control_type_name"},
            "FuelSystemDesign": {"model": "FuelSystemDesign", "attribute": "fuel_system_design_name"},
            "TransmissionNumSpeeds": {"model": "TransmissionNumSpeeds", "attribute": "transmission_num_speeds"},
            "TransmissionControlType": {"model": "TransmissionControlType",
                                        "attribute": "transmission_control_type_name"},
            "TransmissionMfrCode": {"model": "TransmissionMfrCode", "attribute": "transmission_mfr_code"},
            "TransmissionType": {"model": "TransmissionType", "attribute": "transmission_type_name"},
            "TransmissionMfr": {"model": "Mfr", "attribute": "mfr_name"},
            "TransmissionElecControlled": {"model": "ElecControlled", "attribute": "elec_controlled"},
            "EngineMfr": {"model": "EngineConfig2", "attribute": "engine_mfr.mfr_name"},
            "EngineValves": {"model": "EngineConfig2", "attribute": "valves.valves_per_engine"},
            "PowerOutput": {"model": "EngineConfig2", "attribute": "power_output.horse_power"},
            "Region": {"model": "Region", "attribute": "region_name"},
            "PartTerminologyIDs": {"model": None, "attribute": None},
            "QualifierIDs": {"model": None, "attribute": None},
            "Validation Errors": {"model": None, "attribute": None}
        }

        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(default_aces_template, f, indent=2)
                logger.info(f"Created default ACES template at {path}")
        except Exception as e:
            logger.error(f"Error creating default ACES template: {str(e)}")

    def get_template_names(self) -> List[str]:
        """
        Get a list of available template names.

        Returns:
            List of template names
        """
        return list(self.templates.keys())

    def get_template(self, template_name: str) -> Optional[Dict[str, FieldMapping]]:
        """
        Get a specific template by name.

        Args:
            template_name: Name of the template to retrieve

        Returns:
            Template mapping dictionary or None if template doesn't exist
        """
        return self.templates.get(template_name)

    def map_data_to_template(
            self,
            data: List[Dict[str, Any]],
            template_name: str
    ) -> List[Dict[str, Any]]:
        """
        Map the provided data to the specified template format.

        Args:
            data: List of data records to map
            template_name: Name of the template to use for mapping

        Returns:
            List of records mapped to the template format
        """
        template = self.get_template(template_name)
        if not template:
            logger.error(f"Template {template_name} not found")
            return []

        mapped_data = []
        for record in data:
            mapped_record = {}
            for field_name, mapping in template.items():
                field_value = ""

                if mapping["model"] is None or mapping["attribute"] is None:
                    mapped_record[field_name] = ""
                    continue

                # Try different ways to access the data
                model = mapping["model"]
                attribute = mapping["attribute"]

                # 1. Direct column match (e.g., "year_id")
                if attribute in record:
                    field_value = record[attribute]

                # 2. Combined column match (e.g., "Make Make Id")
                elif f"{model} {attribute.replace('_', ' ').title()}" in record:
                    field_value = record[f"{model} {attribute.replace('_', ' ').title()}"]

                # 3. Model name match (e.g., "Make")
                elif model in record:
                    if isinstance(record[model], dict) and attribute in record[model]:
                        field_value = record[model][attribute]

                # 4. Handle nested attributes (e.g., "engine_mfr.mfr_name")
                elif "." in attribute:
                    parts = attribute.split(".")
                    parent_attr = parts[0]
                    child_attr = parts[1]

                    # Try direct nested notation
                    key = f"{parent_attr}.{child_attr}"
                    if key in record:
                        field_value = record[key]

                    # Try model.parent_attr key format
                    elif f"{model}.{parent_attr}" in record:
                        parent_value = record[f"{model}.{parent_attr}"]
                        if isinstance(parent_value, dict) and child_attr in parent_value:
                            field_value = parent_value[child_attr]

                    # Try parent_attr directly
                    elif parent_attr in record:
                        parent_value = record[parent_attr]
                        if isinstance(parent_value, dict) and child_attr in parent_value:
                            field_value = parent_value[child_attr]

                # 5. Try model_attribute format (e.g., "make_name")
                elif f"{model.lower()}_{attribute}" in record:
                    field_value = record[f"{model.lower()}_{attribute}"]

                # 6. Try with display name format
                for key in record:
                    if key.lower().endswith(attribute.lower()) or key.lower().endswith(
                            f"{model.lower()}_{attribute.lower()}"):
                        field_value = record[key]
                        break

                mapped_record[field_name] = field_value

            mapped_data.append(mapped_record)

        return mapped_data

    def export_to_template_csv(
            self,
            data: List[Dict[str, Any]],
            template_name: str,
            output_path: str
    ) -> bool:
        """
        Export data to a CSV file using the specified template format.

        Args:
            data: List of data records to export
            template_name: Name of the template to use for export
            output_path: Path where the CSV file should be saved

        Returns:
            True if export was successful, False otherwise
        """
        try:
            template = self.get_template(template_name)
            if not template:
                logger.error(f"Template {template_name} not found")
                return False

            mapped_data = self.map_data_to_template(data, template_name)
            if not mapped_data:
                logger.warning(f"No data to export after mapping to template {template_name}")
                return False

            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                # Get field names from template
                fieldnames = list(template.keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(mapped_data)

            logger.info(f"Exported {len(mapped_data)} records to {output_path} using template {template_name}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to template CSV: {str(e)}")
            return False

    def export_to_template_excel(
            self,
            data: List[Dict[str, Any]],
            template_name: str,
            output_path: str
    ) -> bool:
        """
        Export data to an Excel file using the specified template format.

        Args:
            data: List of data records to export
            template_name: Name of the template to use for export
            output_path: Path where the Excel file should be saved

        Returns:
            True if export was successful, False otherwise
        """
        try:
            template = self.get_template(template_name)
            if not template:
                logger.error(f"Template {template_name} not found")
                return False

            mapped_data = self.map_data_to_template(data, template_name)
            if not mapped_data:
                logger.warning(f"No data to export after mapping to template {template_name}")
                return False

            wb = Workbook()
            ws = wb.active
            ws.title = template_name

            # Write headers
            fieldnames = list(template.keys())
            for col_idx, field in enumerate(fieldnames, 1):
                ws.cell(row=1, column=col_idx, value=field)

            # Write data
            for row_idx, record in enumerate(mapped_data, 2):
                for col_idx, field in enumerate(fieldnames, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(field, ""))

            wb.save(output_path)
            logger.info(f"Exported {len(mapped_data)} records to {output_path} using template {template_name}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to template Excel: {str(e)}")
            return False

    def create_template_copy_string(
            self,
            data: List[Dict[str, Any]],
            template_name: str,
            delimiter: str = "\t"
    ) -> str:
        """
        Create a string representation of the data using the specified template format,
        suitable for copying to clipboard.

        Args:
            data: List of data records to format
            template_name: Name of the template to use for formatting
            delimiter: Delimiter to use between fields (default: tab character)

        Returns:
            Formatted string representation of the data
        """
        try:
            template = self.get_template(template_name)
            if not template:
                logger.error(f"Template {template_name} not found")
                return ""

            mapped_data = self.map_data_to_template(data, template_name)
            if not mapped_data:
                logger.warning(f"No data to format after mapping to template {template_name}")
                return ""

            # Get field names from template
            fieldnames = list(template.keys())

            lines = []
            # Add header
            lines.append(delimiter.join(fieldnames))

            # Add data rows
            for record in mapped_data:
                row = []
                for field in fieldnames:
                    cell_value = record.get(field, "")
                    # Convert non-string values to strings
                    if not isinstance(cell_value, str):
                        cell_value = str(cell_value) if cell_value is not None else ""
                    row.append(cell_value)
                lines.append(delimiter.join(row))

            return "\n".join(lines)

        except Exception as e:
            logger.error(f"Error creating template copy string: {str(e)}")
            return ""