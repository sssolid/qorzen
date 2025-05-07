from __future__ import annotations

"""
Results panel component for the InitialDB application.

This module provides a panel for displaying query results in a table,
with enhanced options for sorting, filtering, and exporting the data.
"""

import csv
import os
from typing import Any, Dict, List, Optional, Set, Tuple
import structlog
from PyQt6.QtCore import (
    QModelIndex, Qt, pyqtSignal, QSortFilterProxyModel
)
from PyQt6.QtGui import QAction, QKeySequence
from PyQt6.QtWidgets import (
    QApplication, QFileDialog, QFrame, QHBoxLayout, QHeaderView,
    QLabel, QMenu, QMessageBox, QPushButton, QTableView, QVBoxLayout,
    QWidget, QLineEdit, QToolButton, QStatusBar, QSplitter, QSizePolicy
)
from qasync import asyncSlot

try:
    import openpyxl

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from initialdb.services.vehicle_service import VehicleService
from initialdb.utils.dependency_container import resolve
from initialdb.utils.schema_registry import SchemaRegistry

from initialdb.config.settings import EXPORTS_DIR, settings
from initialdb.models.schema import FilterDTO
from initialdb.models.table_model import SortableVehicleTableModel, VehicleResultsTableModel
from initialdb.ui.results_panel.column_selection_dialog import ColumnSelectionDialog
from initialdb.ui.results_panel.export_dialog import ExportDialog
from initialdb.ui.results_panel.multi_filter_panel import MultiFilterPanel
from initialdb.utils.template_manager import TemplateManager
from initialdb.ui.results_panel import ExportHelper

logger = structlog.get_logger(__name__)


class ResultsFilterProxyModel(QSortFilterProxyModel):
    """
    Filter proxy model for results table.

    This model provides filtering capabilities for the results table,
    allowing users to filter by text or column values.
    """

    def __init__(self, parent=None):
        """
        Initialize the filter proxy model.

        Args:
            parent: Parent widget
        """
        super().__init__(parent)
        self._filter_text = ''
        self._filter_columns = set()
        self._column_filters: Dict[str, Any] = {}

    def set_filter_text(self, text: str) -> None:
        """
        Set the text filter.

        Args:
            text: Filter text
        """
        self._filter_text = text.lower()
        self.invalidateFilter()

    def set_filter_columns(self, columns: Set[int]) -> None:
        """
        Set the columns to filter.

        Args:
            columns: Set of column indices
        """
        self._filter_columns = columns
        self.invalidateFilter()

    def set_column_filter(self, column_name: str, filter_value: Any) -> None:
        """
        Set a filter for a specific column.

        Args:
            column_name: Column name
            filter_value: Filter value
        """
        self._column_filters[column_name] = filter_value
        self.invalidateFilter()

    def remove_column_filter(self, column_name: str) -> None:
        """
        Remove a column filter.

        Args:
            column_name: Column name
        """
        if column_name in self._column_filters:
            del self._column_filters[column_name]
            self.invalidateFilter()

    def clear_column_filters(self) -> None:
        """Clear all column filters."""
        self._column_filters.clear()
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        """
        Determine if a row should be accepted by the filter.

        Args:
            source_row: Row index in the source model
            source_parent: Parent index

        Returns:
            True if the row should be accepted
        """
        # Apply column filters
        if self._column_filters:
            for column_name, filter_value in self._column_filters.items():
                column_idx = -1
                header_data = self.sourceModel().headerData

                for col in range(self.sourceModel().columnCount()):
                    if header_data(col, Qt.Orientation.Horizontal) == column_name:
                        column_idx = col
                        break

                if column_idx >= 0:
                    idx = self.sourceModel().index(source_row, column_idx, source_parent)
                    cell_data = self.sourceModel().data(idx, Qt.ItemDataRole.DisplayRole)

                    if isinstance(filter_value, str):
                        if not cell_data or filter_value.lower() not in str(cell_data).lower():
                            return False
                    elif isinstance(filter_value, tuple) and len(filter_value) == 2:
                        if isinstance(filter_value[0], (int, float)) and isinstance(filter_value[1], (int, float)):
                            from_val, to_val = filter_value
                            try:
                                num_value = float(cell_data) if cell_data else 0
                                if num_value < from_val or num_value > to_val:
                                    return False
                            except (ValueError, TypeError):
                                return False
                        elif isinstance(filter_value[0], str) and isinstance(filter_value[1], (int, float)):
                            operator, value = filter_value
                            try:
                                cell_value = float(cell_data) if cell_data else 0
                                if operator == '=' and cell_value != value:
                                    return False
                                elif operator == '>' and cell_value <= value:
                                    return False
                                elif operator == '<' and cell_value >= value:
                                    return False
                                elif operator == '>=' and cell_value < value:
                                    return False
                                elif operator == '<=' and cell_value > value:
                                    return False
                                elif operator == '≠' and cell_value == value:
                                    return False
                            except (ValueError, TypeError):
                                return False

        # Apply text filter
        if not self._filter_text:
            return True

        filter_columns = self._filter_columns or set(range(self.sourceModel().columnCount()))

        for column in filter_columns:
            idx = self.sourceModel().index(source_row, column, source_parent)
            data = self.sourceModel().data(idx, Qt.ItemDataRole.DisplayRole)
            if data and self._filter_text in str(data).lower():
                return True

        return False


class ResultsPanel(QWidget):
    """
    Panel for displaying query results.

    This panel displays the results of vehicle queries in a table,
    with options for filtering, sorting, and exporting the data.
    """

    columnSelectionChanged = pyqtSignal(list)
    resultCountChanged = pyqtSignal(int)

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        """
        Initialize the results panel.

        Args:
            parent: Parent widget
        """
        super().__init__(parent)

        self._registry = resolve(SchemaRegistry)
        self._vehicle_service = resolve(VehicleService)

        self.all_display_fields = self._registry.get_available_display_fields()

        # Initialize with default visible columns
        self.visible_columns = [
            field for field in self.all_display_fields
            if field[1] in ['year_id', 'model_id', 'sub_model_id', 'engine_liter', 'engine_cylinders', 'vehicle_id']
        ]

        # Ensure vehicle_id is included
        if not any((col[1] == 'vehicle_id' for col in self.visible_columns)):
            vehicle_id_field = next((field for field in self.all_display_fields if field[1] == 'vehicle_id'), None)
            if vehicle_id_field:
                self.visible_columns.append(vehicle_id_field)

        # Set up models
        self.results_model = VehicleResultsTableModel()
        self.proxy_model = SortableVehicleTableModel()
        self.proxy_model.setSourceModel(self.results_model)
        self.filter_model = ResultsFilterProxyModel()
        self.filter_model.setSourceModel(self.proxy_model)

        self.template_manager = TemplateManager()

        self._init_ui()

    def _init_ui(self) -> None:
        """Initialize the UI components."""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)

        self.splitter = QSplitter(Qt.Orientation.Vertical)

        # Top panel with filters
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setContentsMargins(2, 2, 2, 2)

        # Header with search and controls
        header_frame = QFrame()
        header_frame.setFrameShape(QFrame.Shape.StyledPanel)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(4, 4, 4, 4)

        self.results_count_label = QLabel('No results')
        header_layout.addWidget(self.results_count_label)

        header_layout.addWidget(QLabel('Search:'))

        self.search_field = QLineEdit()
        self.search_field.setPlaceholderText('Quick filter...')
        self.search_field.textChanged.connect(self._filter_results)
        header_layout.addWidget(self.search_field)

        clear_filter_btn = QToolButton()
        clear_filter_btn.setText('✕')
        clear_filter_btn.setToolTip('Clear quick filter')
        clear_filter_btn.clicked.connect(self._clear_filter)
        header_layout.addWidget(clear_filter_btn)

        select_columns_btn = QPushButton('Select Columns')
        select_columns_btn.clicked.connect(self._show_column_selection_dialog)
        header_layout.addWidget(select_columns_btn)

        export_btn = QPushButton('Export')
        export_btn.setMenu(self._create_export_menu())
        header_layout.addWidget(export_btn)

        top_layout.addWidget(header_frame)

        # Filter panel for additional filtering
        self.filter_panel = MultiFilterPanel()
        self.filter_panel.filtersChanged.connect(self._apply_column_filters)
        self.filter_panel.setMinimumHeight(100)
        self.filter_panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)

        filter_frame = QFrame()
        filter_frame.setFrameShape(QFrame.Shape.StyledPanel)
        filter_layout = QVBoxLayout(filter_frame)
        filter_layout.addWidget(self.filter_panel)

        top_layout.addWidget(filter_frame)
        self.splitter.addWidget(top_widget)

        # Results table
        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        table_layout.setContentsMargins(2, 2, 2, 2)

        self.results_table = QTableView()
        self.results_table.setModel(self.filter_model)
        self.results_table.setSortingEnabled(True)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.results_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.results_table.customContextMenuRequested.connect(self._show_context_menu)

        header = self.results_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setStretchLastSection(True)

        # Add copy action
        copy_action = QAction('Copy', self)
        copy_action.setShortcut(QKeySequence.StandardKey.Copy)
        copy_action.triggered.connect(self._copy_selection)
        self.results_table.addAction(copy_action)

        table_layout.addWidget(self.results_table)

        self.status_bar = QStatusBar()
        self.status_bar.setSizeGripEnabled(False)
        table_layout.addWidget(self.status_bar)

        self.splitter.addWidget(table_widget)
        self.splitter.setSizes([150, 650])

        main_layout.addWidget(self.splitter)

    def _filter_results(self, text: str) -> None:
        """
        Filter results by text.

        Args:
            text: Filter text
        """
        self.filter_model.set_filter_text(text)

        visible_count = self.filter_model.rowCount()
        total_count = self.results_model.rowCount()

        if text:
            self.status_bar.showMessage(f'Showing {visible_count} of {total_count} results')
        else:
            self.status_bar.showMessage('')

        self.resultCountChanged.emit(visible_count)

    def _apply_column_filters(self, filters: Dict[str, Any]) -> None:
        """
        Apply column filters.

        Args:
            filters: Dictionary of column filters
        """
        self.filter_model.clear_column_filters()

        for column_key, filter_value in filters.items():
            if '.' in column_key:
                parts = column_key.split('.', 1)
                display_name = parts[1]

                for _, col, disp in self.all_display_fields:
                    if col == parts[1]:
                        display_name = disp
                        break

                self.filter_model.set_column_filter(display_name, filter_value)
            else:
                self.filter_model.set_column_filter(column_key, filter_value)

        visible_count = self.filter_model.rowCount()
        total_count = self.results_model.rowCount()

        if visible_count < total_count:
            self.status_bar.showMessage(f'Showing {visible_count} of {total_count} results')
        else:
            self.status_bar.showMessage('')

        self.resultCountChanged.emit(visible_count)

    def _clear_filter(self) -> None:
        """Clear the text filter."""
        self.search_field.clear()

    def _create_export_menu(self) -> QMenu:
        """
        Create the export menu.

        Returns:
            Export menu
        """
        menu = QMenu(self)

        # Standard export options
        standard_menu = QMenu('Standard Export', self)

        csv_action = QAction('Export to CSV', self)
        csv_action.triggered.connect(lambda: self._export_data('csv'))
        standard_menu.addAction(csv_action)

        excel_action = QAction('Export to Excel', self)
        excel_action.triggered.connect(lambda: self._export_data('excel'))
        excel_action.setEnabled(EXCEL_AVAILABLE)
        standard_menu.addAction(excel_action)

        csv_selected_action = QAction('Export Selected Rows to CSV', self)
        csv_selected_action.triggered.connect(lambda: self._export_data('csv', selected_only=True))
        standard_menu.addAction(csv_selected_action)

        excel_selected_action = QAction('Export Selected Rows to Excel', self)
        excel_selected_action.triggered.connect(lambda: self._export_data('excel', selected_only=True))
        excel_selected_action.setEnabled(EXCEL_AVAILABLE)
        standard_menu.addAction(excel_selected_action)

        menu.addMenu(standard_menu)

        # Template export options
        templates_menu = QMenu('Template Export', self)
        template_names = self.template_manager.get_template_names()

        if template_names:
            for template_name in template_names:
                template_submenu = QMenu(template_name, self)

                csv_template_action = QAction('Export to CSV', self)
                csv_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('csv', t)
                )
                template_submenu.addAction(csv_template_action)

                excel_template_action = QAction('Export to Excel', self)
                excel_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('excel', t)
                )
                excel_template_action.setEnabled(EXCEL_AVAILABLE)
                template_submenu.addAction(excel_template_action)

                template_submenu.addSeparator()

                csv_selected_template_action = QAction('Export Selected Rows to CSV', self)
                csv_selected_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('csv', t, True)
                )
                template_submenu.addAction(csv_selected_template_action)

                excel_selected_template_action = QAction('Export Selected Rows to Excel', self)
                excel_selected_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('excel', t, True)
                )
                excel_selected_template_action.setEnabled(EXCEL_AVAILABLE)
                template_submenu.addAction(excel_selected_template_action)

                templates_menu.addMenu(template_submenu)
        else:
            no_templates_action = QAction('No templates available', self)
            no_templates_action.setEnabled(False)
            templates_menu.addAction(no_templates_action)

        menu.addMenu(templates_menu)

        return menu

    def _show_context_menu(self, pos) -> None:
        """
        Show the context menu for the results table.

        Args:
            pos: Position to show the menu
        """
        menu = QMenu(self)

        # Copy actions
        copy_action = QAction('Copy', self)
        copy_action.triggered.connect(self._copy_selection)
        menu.addAction(copy_action)

        copy_with_headers_action = QAction('Copy with Headers', self)
        copy_with_headers_action.triggered.connect(lambda: self._copy_selection(with_headers=True))
        menu.addAction(copy_with_headers_action)

        # Template copy actions
        if self.template_manager.get_template_names():
            templates_copy_menu = QMenu('Copy to Template', self)

            for template_name in self.template_manager.get_template_names():
                template_copy_action = QAction(template_name, self)
                template_copy_action.triggered.connect(
                    lambda checked, t=template_name: self._copy_to_template(t)
                )
                templates_copy_menu.addAction(template_copy_action)

            menu.addMenu(templates_copy_menu)

        menu.addSeparator()

        # Export actions
        export_menu = menu.addMenu('Export')

        csv_action = QAction('Export to CSV', self)
        csv_action.triggered.connect(lambda: self._export_data('csv'))
        export_menu.addAction(csv_action)

        excel_action = QAction('Export to Excel', self)
        excel_action.triggered.connect(lambda: self._export_data('excel'))
        excel_action.setEnabled(EXCEL_AVAILABLE)
        export_menu.addAction(excel_action)

        csv_selected_action = QAction('Export Selected Rows to CSV', self)
        csv_selected_action.triggered.connect(lambda: self._export_data('csv', selected_only=True))
        export_menu.addAction(csv_selected_action)

        excel_selected_action = QAction('Export Selected Rows to Excel', self)
        excel_selected_action.triggered.connect(lambda: self._export_data('excel', selected_only=True))
        excel_selected_action.setEnabled(EXCEL_AVAILABLE)
        export_menu.addAction(excel_selected_action)

        # Template export actions
        if self.template_manager.get_template_names():
            export_menu.addSeparator()
            templates_menu = export_menu.addMenu('Template Export')

            for template_name in self.template_manager.get_template_names():
                template_submenu = QMenu(template_name, self)

                csv_template_action = QAction('Export to CSV', self)
                csv_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('csv', t)
                )
                template_submenu.addAction(csv_template_action)

                excel_template_action = QAction('Export to Excel', self)
                excel_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('excel', t)
                )
                excel_template_action.setEnabled(EXCEL_AVAILABLE)
                template_submenu.addAction(excel_template_action)

                template_submenu.addSeparator()

                csv_selected_template_action = QAction('Export Selected Rows to CSV', self)
                csv_selected_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('csv', t, True)
                )
                template_submenu.addAction(csv_selected_template_action)

                excel_selected_template_action = QAction('Export Selected Rows to Excel', self)
                excel_selected_template_action.triggered.connect(
                    lambda checked, t=template_name: self._export_to_template('excel', t, True)
                )
                excel_selected_template_action.setEnabled(EXCEL_AVAILABLE)
                template_submenu.addAction(excel_selected_template_action)

                templates_menu.addMenu(template_submenu)

        menu.addSeparator()

        # Selection actions
        select_all_action = QAction('Select All', self)
        select_all_action.triggered.connect(self.results_table.selectAll)
        menu.addAction(select_all_action)

        menu.exec(self.results_table.viewport().mapToGlobal(pos))

    def _copy_selection(self, with_headers: bool = False) -> None:
        """
        Copy selected cells to clipboard.

        Args:
            with_headers: Whether to include headers
        """
        selection = self.results_table.selectionModel()
        if not selection.hasSelection():
            return

        indexes = selection.selectedIndexes()
        if not indexes:
            return

        model = self.results_table.model()
        rows = set((index.row() for index in indexes))
        columns = set((index.column() for index in indexes))

        rows_list = sorted(rows)
        columns_list = sorted(columns)

        lines = []

        # Add headers if requested
        if with_headers:
            header_row = []
            for col in columns_list:
                header_text = model.headerData(col, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                header_row.append(str(header_text) if header_text else '')
            lines.append('\t'.join(header_row))

        # Add data rows
        for row in rows_list:
            cells = []
            for col in columns_list:
                index = model.index(row, col)
                data = model.data(index, Qt.ItemDataRole.DisplayRole)
                cells.append(str(data) if data else '')
            lines.append('\t'.join(cells))

        text = '\n'.join(lines)
        QApplication.clipboard().setText(text)

    def _copy_to_template(self, template_name: str) -> None:
        """
        Copy selected rows using a template format.

        Args:
            template_name: Name of the template to use
        """
        selection = self.results_table.selectionModel()
        if not selection.hasSelection():
            QMessageBox.information(self, 'Copy to Template', 'No rows selected to copy.')
            return

        selected_rows = sorted(set((index.row() for index in selection.selectedIndexes())))

        export_helper = ExportHelper(self.results_table.model(), self.visible_columns)
        rows_data = export_helper._get_selected_rows_data(selected_rows=selected_rows, required_columns=None)

        if not rows_data:
            QMessageBox.information(self, 'Copy to Template', 'No data to copy.')
            return

        template_string = self.template_manager.create_template_copy_string(
            data=rows_data, template_name=template_name
        )

        if template_string:
            QApplication.clipboard().setText(template_string)
            self.status_bar.showMessage(
                f'Copied {len(rows_data)} rows to clipboard using {template_name} template format',
                5000
            )
        else:
            QMessageBox.warning(
                self, 'Copy to Template',
                'Failed to create template copy. Please check the template structure.'
            )
            self.status_bar.showMessage('Failed to copy data to template format', 5000)

    def _export_data(self, format_type: str, selected_only: bool = False) -> None:
        """
        Export data to a file.

        Args:
            format_type: 'csv' or 'excel'
            selected_only: Whether to export only selected rows
        """
        if format_type == 'excel' and (not EXCEL_AVAILABLE):
            QMessageBox.warning(
                self, 'Export Error',
                'Excel export requires the openpyxl package. Please install it with pip install openpyxl.'
            )
            return

        model = self.results_table.model()
        if model.rowCount() == 0:
            QMessageBox.information(self, 'Export', 'No data available to export.')
            return

        has_selection = False
        if selected_only:
            selection = self.results_table.selectionModel()
            has_selection = selection.hasSelection()
            if not has_selection:
                QMessageBox.information(self, 'Export', 'No rows selected for export.')
                return

        # Show export dialog
        dialog = ExportDialog(
            has_excel=EXCEL_AVAILABLE,
            has_selection=self.results_table.selectionModel().hasSelection(),
            parent=self
        )

        if format_type == 'excel':
            dialog.excel_radio.setChecked(True)

        if selected_only:
            dialog.selected_check.setChecked(True)

        if dialog.exec() != ExportDialog.DialogCode.Accepted:
            return

        format_type, filename, selected_only, include_headers, template_name = dialog.get_export_options()

        if template_name:
            self._export_to_template(format_type, template_name, selected_only, filename)
            return

        # Get headers
        headers = []
        if include_headers:
            headers = [
                model.headerData(col, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                for col in range(model.columnCount())
            ]

        # Get rows to export
        rows_to_export = []
        if selected_only:
            selection = self.results_table.selectionModel()
            rows_to_export = list(set((index.row() for index in selection.selectedIndexes())))
            rows_to_export.sort()
        else:
            rows_to_export = list(range(model.rowCount()))

        # Export data
        try:
            if format_type == 'csv':
                self._export_to_csv(filename, headers if include_headers else [], model, rows_to_export)
            else:
                self._export_to_excel(filename, headers if include_headers else [], model, rows_to_export)

            settings.add_recent_export(filename)
            QMessageBox.information(self, 'Export Successful', f'Data exported successfully to {filename}')
        except Exception as e:
            logger.error(f'Export error: {str(e)}')
            QMessageBox.critical(self, 'Export Error', f'Error exporting data: {str(e)}')

    def _export_to_template(self, format_type: str, template_name: str,
                            selected_only: bool = False, filename: Optional[str] = None) -> None:
        """
        Export data using a template.

        Args:
            format_type: 'csv' or 'excel'
            template_name: Name of the template to use
            selected_only: Whether to export only selected rows
            filename: Optional filename to use
        """
        if format_type == 'excel' and (not EXCEL_AVAILABLE):
            QMessageBox.warning(
                self, 'Export Error',
                'Excel export requires the openpyxl package. Please install it with pip install openpyxl.'
            )
            return

        # Get selected rows if needed
        selected_rows = None
        if selected_only:
            selection = self.results_table.selectionModel()
            if not selection.hasSelection():
                QMessageBox.information(self, 'Export', 'No rows selected for export.')
                return
            selected_rows = sorted(set((index.row() for index in selection.selectedIndexes())))

        # Show status
        original_status = self.status_bar.currentMessage()
        self.status_bar.showMessage('Preparing export data...')
        QApplication.processEvents()

        # Create export helper
        export_helper = ExportHelper(self.results_table.model(), self.visible_columns, repository=None)

        # Get filename if not provided
        if not filename:
            export_dir = settings.get('default_exports_path', str(EXPORTS_DIR))
            if not os.path.exists(export_dir):
                os.makedirs(export_dir, exist_ok=True)

            base_name = f'vehicle_query_results_{template_name}'
            extension = '.xlsx' if format_type == 'excel' else '.csv'

            filename, _ = QFileDialog.getSaveFileName(
                self, 'Export Data',
                os.path.join(export_dir, f'{base_name}{extension}'),
                'CSV Files (*.csv)' if format_type == 'csv' else 'Excel Files (*.xlsx)'
            )

            if not filename:
                self.status_bar.showMessage(original_status)
                return

        # Export data
        self.status_bar.showMessage(f'Exporting to {template_name} template...')
        QApplication.processEvents()

        success = export_helper.export_to_template(
            format_type=format_type,
            template_name=template_name,
            selected_only=selected_only,
            selected_rows=selected_rows,
            filename=filename
        )

        if success:
            settings.add_recent_export(filename)
            self.status_bar.showMessage(f'Export successful: {filename}', 5000)
            QMessageBox.information(
                self, 'Export Successful',
                f'Data exported successfully to {filename} using {template_name} template'
            )
        else:
            self.status_bar.showMessage('Export failed', 5000)
            QMessageBox.critical(
                self, 'Export Error',
                f'Failed to export data using {template_name} template'
            )

    def _export_to_csv(self, filename: str, headers: List[str], model: Any, rows: List[int]) -> None:
        """
        Export data to CSV.

        Args:
            filename: Output filename
            headers: Column headers
            model: Data model
            rows: Rows to export
        """
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            if headers:
                writer.writerow(headers)

            for row in rows:
                row_data = []
                for col in range(model.columnCount()):
                    index = model.index(row, col)
                    data = model.data(index, Qt.ItemDataRole.DisplayRole)
                    row_data.append(data if data else '')
                writer.writerow(row_data)

    def _export_to_excel(self, filename: str, headers: List[str], model: Any, rows: List[int]) -> None:
        """
        Export data to Excel.

        Args:
            filename: Output filename
            headers: Column headers
            model: Data model
            rows: Rows to export
        """
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vehicle Query Results'

        start_row = 1

        # Write headers
        if headers:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            start_row = 2

        # Write data
        for row_idx, row in enumerate(rows, start_row):
            for col_idx in range(1, model.columnCount() + 1):
                index = model.index(row, col_idx - 1)
                data = model.data(index, Qt.ItemDataRole.DisplayRole)
                ws.cell(row=row_idx, column=col_idx, value=data if data else '')

        # Adjust column widths
        for col_idx in range(1, model.columnCount() + 1):
            column_letter = get_column_letter(col_idx)
            if headers:
                header = headers[col_idx - 1] if col_idx <= len(headers) else ''
                ws.column_dimensions[column_letter].width = max(10, len(str(header)) + 2)
            else:
                ws.column_dimensions[column_letter].width = 12

        wb.save(filename)

    def _get_all_rows_data(self) -> List[Dict[str, Any]]:
        """
        Get all rows of data from the model.

        Returns:
            List of dictionaries with row data
        """
        model = self.results_table.model()
        source_model = None

        if hasattr(model, 'sourceModel'):
            source_model = model.sourceModel()
            if hasattr(source_model, 'sourceModel'):
                source_model = source_model.sourceModel()

        if source_model and hasattr(source_model, 'getRawData'):
            return source_model.getRawData()

        raw_data = []
        for row in range(model.rowCount()):
            row_data = {}
            for col in range(model.columnCount()):
                header = model.headerData(col, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                if header:
                    index = model.index(row, col)
                    value = model.data(index, Qt.ItemDataRole.DisplayRole)
                    row_data[str(header)] = value
            raw_data.append(row_data)

        return raw_data

    def _get_selected_rows_data(self) -> List[Dict[str, Any]]:
        """
        Get data for selected rows.

        Returns:
            List of dictionaries with row data
        """
        selection = self.results_table.selectionModel()
        if not selection.hasSelection():
            return []

        model = self.results_table.model()
        rows = sorted(set((index.row() for index in selection.selectedIndexes())))

        source_model = None
        if hasattr(model, 'sourceModel'):
            source_model = model.sourceModel()
            if hasattr(source_model, 'sourceModel'):
                source_model = source_model.sourceModel()

        if source_model and hasattr(source_model, 'getRawData'):
            all_raw_data = source_model.getRawData()
            selected_data = []

            for row in rows:
                source_row = row
                temp_model = model
                while hasattr(temp_model, 'mapToSource'):
                    source_index = temp_model.mapToSource(temp_model.index(source_row, 0))
                    source_row = source_index.row()
                    temp_model = temp_model.sourceModel()

                if 0 <= source_row < len(all_raw_data):
                    selected_data.append(all_raw_data[source_row])

            return selected_data

        raw_data = []
        for row in rows:
            row_data = {}
            for col in range(model.columnCount()):
                header = model.headerData(col, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                if header:
                    index = model.index(row, col)
                    value = model.data(index, Qt.ItemDataRole.DisplayRole)
                    row_data[str(header)] = value
            raw_data.append(row_data)

        return raw_data

    def _show_column_selection_dialog(self) -> None:
        """Show dialog for selecting which columns to display."""
        # Get the available filters from the registry
        available_filters = self._registry.get_available_filters()

        # Use only columns that match the available filters
        available_filter_keys = {(table, column) for table, column, _ in available_filters}
        filtered_columns = [
            col for col in self.all_display_fields
            if (col[0], col[1]) in available_filter_keys
        ]

        dialog = ColumnSelectionDialog(filtered_columns, self.visible_columns, self)
        if dialog.exec() == ColumnSelectionDialog.DialogCode.Accepted:
            self.visible_columns = dialog.get_visible_columns()

            # Ensure vehicle_id is included
            if not any((col[1] == 'vehicle_id' for col in self.visible_columns)):
                vehicle_id_field = next(
                    (field for field in self.all_display_fields if field[1] == 'vehicle_id'),
                    None
                )
                if vehicle_id_field:
                    self.visible_columns.append(vehicle_id_field)

            self.columnSelectionChanged.emit(self.visible_columns)
            self._apply_column_visibility()

    def _apply_column_visibility(self) -> None:
        """Apply column visibility based on visible_columns."""
        if self.results_model.rowCount() == 0:
            return

        current_headers = self.results_model.getHeaders()
        logger.debug(f'Current headers: {current_headers}')

        aliases = [alias for _, _, alias in self.visible_columns]
        if 'vehicle_id' not in aliases:
            aliases.append('vehicle_id')

        logger.debug(f'Looking for these column aliases: {aliases}')

        for col_idx, header in enumerate(current_headers):
            is_visible = header in aliases
            logger.debug(f"Column '{header}' (idx {col_idx}) visible: {is_visible}")
            self.results_table.setColumnHidden(col_idx, not is_visible)

        self.results_table.resizeColumnsToContents()

    @asyncSlot()
    async def execute_query(self, filters: FilterDTO) -> None:
        """
        Execute a query and display results.

        Args:
            filters: Filter criteria
        """
        visible_columns = self.get_visible_columns()
        if not any((col[1] == 'vehicle_id' for col in visible_columns)):
            visible_columns.append(('vehicle', 'vehicle_id', 'vehicle_id'))

        self.status_bar.showMessage('Executing query...')
        self.clear_results()

        try:
            results = await self._vehicle_service.get_vehicles(filters, visible_columns)
            self.set_results(results)

            count = len(results)
            self.status_bar.showMessage(f'Query complete: {count} results')
        except Exception as e:
            logger.error(f'Error executing query: {str(e)}')
            self.status_bar.showMessage(f'Query error: {str(e)}')
            QMessageBox.critical(self, 'Query Error', f'Error executing query: {str(e)}')

    def set_results(self, results: List[Dict[str, Any]]) -> None:
        """
        Set query results.

        Args:
            results: Query results as a list of dictionaries
        """
        if isinstance(results, dict):
            # Unwrap the first (and presumably only) key of the dict
            first_key = next(iter(results))
            data = results[first_key]
        else:
            data = results

        logger.debug(f'Setting results data with {len(data)} rows')
        self.results_model.setData(data)

        result_count = len(results)
        self.results_count_label.setText(f"{result_count} result{('s' if result_count != 1 else '')}")

        self._apply_column_visibility()
        self._clear_filter()

        filtered_count = self.filter_model.rowCount()
        if filtered_count != result_count:
            self.status_bar.showMessage(f'Showing {filtered_count} of {result_count} results')
        else:
            self.status_bar.showMessage('')

        self.resultCountChanged.emit(filtered_count)

    def get_visible_columns(self) -> List[Tuple[str, str, str]]:
        """
        Get visible columns.

        Returns:
            List of visible columns
        """
        return self.visible_columns

    def set_visible_columns(self, columns: List[Tuple[str, str, str]]) -> None:
        """
        Set visible columns.

        Args:
            columns: List of columns to make visible
        """
        self.visible_columns = columns

        # Ensure vehicle_id is included
        if not any((col[1] == 'vehicle_id' for col in self.visible_columns)):
            vehicle_id_field = next((field for field in self.all_display_fields if field[1] == 'vehicle_id'), None)
            if vehicle_id_field:
                self.visible_columns.append(vehicle_id_field)

        self._apply_column_visibility()

    def clear_results(self) -> None:
        """Clear all results."""
        self.results_model.clearData()
        self.results_count_label.setText('No results')
        self.status_bar.showMessage('')
        self._clear_filter()
        self.filter_panel.clear_all_filters()
        self.resultCountChanged.emit(0)

    def show_export_dialog(self) -> None:
        """Show the export dialog."""
        self._export_data('csv')

    def show_column_selection_dialog(self) -> None:
        """Show the column selection dialog."""
        self._show_column_selection_dialog()