#!/usr/bin/env python3
# export_service.py
from __future__ import annotations

"""
Export service for exporting vehicle data to various formats.

This module provides functionality for exporting vehicle data to CSV, Excel,
and other formats, with support for templates and custom formatting.
"""

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, cast

from ..models.schema import FilterDTO
from ..exceptions import ExportError
from ..config.settings import get_plugin_config_namespace


class ExportService:
    """Service for exporting vehicle data to various formats.

    Provides methods to export data to CSV, Excel, or other formats
    using configurable templates.
    """

    def __init__(self, file_manager: Any, logger: Any, vehicle_service: Any, config: Any) -> None:
        """Initialize the export service.

        Args:
            file_manager: File manager from Qorzen
            logger: Logger for this service
            vehicle_service: Vehicle service instance
            config: Config provider from Qorzen
        """
        self._file_manager = file_manager
        self._logger = logger
        self._vehicle_service = vehicle_service
        self._config = config

        # Get configuration using Qorzen's config manager
        namespace = get_plugin_config_namespace()
        self._exports_dir = self._config.get(f"{namespace}.exports_dir", "exports")
        self._templates_dir = self._config.get(f"{namespace}.templates_dir", "templates")

        # Ensure directories exist
        self._ensure_directories()

        # Load templates
        self._templates = self._load_templates()

        self._logger.info("Export service initialized")

    def _ensure_directories(self) -> None:
        """Ensure export and template directories exist."""
        try:
            self._file_manager.ensure_directory(self._exports_dir, directory_type="plugin_data")
            self._file_manager.ensure_directory(self._templates_dir, directory_type="plugin_data")
            self._logger.debug(f"Directories ensured: {self._exports_dir}, {self._templates_dir}")
        except Exception as e:
            self._logger.error(f"Failed to create directories: {e}")

    def _load_templates(self) -> Dict[str, Dict[str, Any]]:
        """Load export templates from template directory.

        Returns:
            Dictionary of template name to template definition
        """
        templates = {}
        try:
            template_files = self._file_manager.list_files(
                path=self._templates_dir,
                directory_type="plugin_data",
                pattern="*.json"
            )

            for template_file in template_files:
                try:
                    content = self._file_manager.read_text(
                        template_file.path,
                        directory_type="plugin_data"
                    )
                    template_data = json.loads(content)
                    template_name = Path(template_file.path).stem
                    templates[template_name] = template_data
                    self._logger.debug(f"Loaded template: {template_name}")
                except Exception as e:
                    self._logger.error(f"Failed to load template {template_file.path}: {e}")

            self._logger.info(f"Loaded {len(templates)} templates")

            # Create default template if none exist
            if not templates:
                self._create_default_template()
                templates = self._load_templates()
        except Exception as e:
            self._logger.error(f"Failed to load templates: {e}")

        return templates

    def _create_default_template(self) -> None:
        """Create default ACES template if none exists."""
        try:
            default_template = {
                "name": "ACES_Template",
                "description": "Default ACES export template",
                "mappings": {
                    "Make": {"field": "make_name"},
                    "Model": {"field": "model_name"},
                    "Year": {"field": "year_id"},
                    "SubModel": {"field": "sub_model_name"},
                    "Region": {"field": "region_name"},
                    "EngineBase": {"field": "liter", "suffix": "L"},
                    "EngineCylinders": {"field": "cylinders"},
                    "FuelType": {"field": "fuel_type_name"},
                    "BodyType": {"field": "body_type_name"}
                }
            }

            template_path = f"{self._templates_dir}/ACES_Template.json"

            self._file_manager.write_text(
                template_path,
                content=json.dumps(default_template, indent=2),
                directory_type="plugin_data"
            )

            self._logger.info("Created default ACES template")
        except Exception as e:
            self._logger.error(f"Failed to create default template: {e}")

    def update_config(self, config: Any) -> None:
        """Update service configuration.

        Args:
            config: Config provider from Qorzen
        """
        self._config = config

        # Update configuration from the provider
        namespace = get_plugin_config_namespace()
        self._exports_dir = self._config.get(f"{namespace}.exports_dir", "exports")
        self._templates_dir = self._config.get(f"{namespace}.templates_dir", "templates")

        # Reload with new paths
        self._ensure_directories()
        self._templates = self._load_templates()

        self._logger.debug(
            f"Configuration updated: exports_dir={self._exports_dir}, templates_dir={self._templates_dir}")

    def get_available_templates(self) -> List[Dict[str, Any]]:
        """Get available export templates.

        Returns:
            List of template metadata
        """
        templates = []
        for name, template in self._templates.items():
            templates.append({
                "name": name,
                "description": template.get("description", "")
            })
        return templates

    def get_template(self, name: str) -> Optional[Dict[str, Any]]:
        """Get template by name."""
        return self._templates.get(name)

    def export_to_csv(self, data: List[Dict[str, Any]], output_path: str,
                      template_name: Optional[str] = None) -> str:
        """Export data to CSV file.

        Args:
            data: List of data records
            output_path: Output file path
            template_name: Template name to apply

        Returns:
            Absolute path to the created file

        Raises:
            ExportError: If export fails
        """
        try:
            # Apply template if specified
            if template_name and template_name in self._templates:
                processed_data = self._apply_template(data, template_name)
                field_names = list(self._templates[template_name]["mappings"].keys())
            else:
                processed_data = data
                field_names = self._get_field_names(data)

            # Ensure the output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                self._file_manager.ensure_directory(output_dir, directory_type="plugin_data")

            # Write CSV file
            with self._file_manager.open_file(output_path, "w", directory_type="plugin_data", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=field_names)
                writer.writeheader()
                writer.writerows(processed_data)

            self._logger.info(f"Exported {len(data)} records to {output_path}")

            return output_path
        except Exception as e:
            self._logger.error(f"Failed to export to CSV: {e}")
            raise ExportError(f"Failed to export to CSV: {e}") from e

    def export_to_excel(self, data: List[Dict[str, Any]], output_path: str,
                        template_name: Optional[str] = None) -> str:
        """Export data to Excel file.

        Args:
            data: List of data records
            output_path: Output file path
            template_name: Template name to apply

        Returns:
            Absolute path to the created file

        Raises:
            ExportError: If export fails
        """
        try:
            # Import openpyxl here to avoid dependencies if Excel export is not used
            from openpyxl import Workbook
        except ImportError:
            raise ExportError("Excel export not supported - openpyxl not installed")

        try:
            # Apply template if specified
            if template_name and template_name in self._templates:
                processed_data = self._apply_template(data, template_name)
                field_names = list(self._templates[template_name]["mappings"].keys())
            else:
                processed_data = data
                field_names = self._get_field_names(data)

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active

            # Write header row
            for col_idx, field in enumerate(field_names, 1):
                ws.cell(row=1, column=col_idx, value=field)

            # Write data rows
            for row_idx, record in enumerate(processed_data, 2):
                for col_idx, field in enumerate(field_names, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(field, ""))

            # Ensure the output directory exists
            output_dir = os.path.dirname(output_path)
            if output_dir:
                self._file_manager.ensure_directory(output_dir, directory_type="plugin_data")

            # Save the workbook
            wb.save(self._file_manager.get_absolute_path(output_path, directory_type="plugin_data"))

            self._logger.info(f"Exported {len(data)} records to Excel file {output_path}")

            return output_path
        except Exception as e:
            self._logger.error(f"Failed to export to Excel: {e}")
            raise ExportError(f"Failed to export to Excel: {e}") from e

    def export_query_results(self, filters: FilterDTO, format: str,
                             template_name: Optional[str] = None,
                             filename: Optional[str] = None) -> str:
        """Export query results based on filters.

        Args:
            filters: Filter criteria for the query
            format: Export format (csv, excel)
            template_name: Template name to apply
            filename: Custom filename (generated if None)

        Returns:
            Path to the exported file

        Raises:
            ExportError: If export fails
        """
        try:
            # Get data from vehicle service
            data = self._vehicle_service.get_vehicles(filters)

            if not data:
                raise ExportError("No data to export")

            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"vehicle_export_{timestamp}.{format.lower()}"

            # Full output path
            output_path = f"{self._exports_dir}/{filename}"

            # Export in the requested format
            if format.lower() == "csv":
                return self.export_to_csv(data, output_path, template_name)
            elif format.lower() == "excel":
                return self.export_to_excel(data, output_path, template_name)
            else:
                raise ExportError(f"Unsupported export format: {format}")
        except Exception as e:
            self._logger.error(f"Failed to export query results: {e}")
            raise ExportError(f"Failed to export query results: {e}") from e

    def _apply_template(self, data: List[Dict[str, Any]], template_name: str) -> List[Dict[str, Any]]:
        """Apply a template to the data.

        Args:
            data: Source data records
            template_name: Template name to apply

        Returns:
            Transformed data records
        """
        template = self._templates[template_name]
        mappings = template["mappings"]
        result = []

        for record in data:
            processed_record = {}

            for output_field, mapping in mappings.items():
                source_field = mapping["field"]
                value = record.get(source_field, "")

                # Apply prefix if specified
                if "prefix" in mapping and value:
                    value = f"{mapping['prefix']}{value}"

                # Apply suffix if specified
                if "suffix" in mapping and value:
                    value = f"{value}{mapping['suffix']}"

                processed_record[output_field] = value

            result.append(processed_record)

        return result

    def _get_field_names(self, data: List[Dict[str, Any]]) -> List[str]:
        """Get field names from data.

        Args:
            data: List of data records

        Returns:
            List of field names
        """
        if not data:
            return []

        # Collect all unique field names
        all_fields = set()
        for record in data:
            all_fields.update(record.keys())

        return sorted(list(all_fields))

    def create_template(self, name: str, description: str, mappings: Dict[str, Dict[str, str]]) -> None:
        """Create a new export template.

        Args:
            name: Template name
            description: Template description
            mappings: Field mappings

        Raises:
            ExportError: If template creation fails
        """
        try:
            template = {
                "name": name,
                "description": description,
                "mappings": mappings
            }

            template_path = f"{self._templates_dir}/{name}.json"

            self._file_manager.write_text(
                template_path,
                content=json.dumps(template, indent=2),
                directory_type="plugin_data"
            )

            # Reload templates
            self._templates = self._load_templates()

            self._logger.info(f"Created template: {name}")
        except Exception as e:
            self._logger.error(f"Failed to create template: {e}")
            raise ExportError(f"Failed to create template: {e}") from e

    def delete_template(self, name: str) -> None:
        """Delete an export template.

        Args:
            name: Template name to delete

        Raises:
            ExportError: If template deletion fails
        """
        try:
            if name not in self._templates:
                raise ExportError(f"Template not found: {name}")

            template_path = f"{self._templates_dir}/{name}.json"

            self._file_manager.delete_file(template_path, directory_type="plugin_data")

            # Reload templates
            self._templates = self._load_templates()

            self._logger.info(f"Deleted template: {name}")
        except Exception as e:
            self._logger.error(f"Failed to delete template: {e}")
            raise ExportError(f"Failed to delete template: {e}") from e

    def get_recent_exports(self, limit: int = 10) -> List[Dict[str, Any]]:
        """Get list of recent exports.

        Args:
            limit: Maximum number of exports to return

        Returns:
            List of export metadata
        """
        try:
            export_files = self._file_manager.list_files(
                path=self._exports_dir,
                directory_type="plugin_data",
                recursive=False
            )

            # Sort by modification time, newest first
            export_files.sort(key=lambda f: f.modified_at, reverse=True)

            exports = []
            for file in export_files[:limit]:
                exports.append({
                    "path": file.path,
                    "filename": os.path.basename(file.path),
                    "timestamp": file.modified_at.isoformat(),
                    "size": file.size
                })

            return exports
        except Exception as e:
            self._logger.error(f"Failed to get recent exports: {e}")
            return []

    def shutdown(self) -> None:
        """Shut down the service."""
        self._logger.info("Export service shutting down")