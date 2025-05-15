from __future__ import annotations

"""
VCdb data export module.

This module provides functionality for exporting query results to different file formats,
including CSV and Excel.
"""

import csv
import logging
import os
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple, Union, Awaitable

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportError(Exception):
    """Exception raised for errors during data export."""

    def __init__(self, message: str, details: Optional[Dict[str, Any]] = None) -> None:
        """
        Initialize the exception.

        Args:
            message: Error message
            details: Additional details about the error
        """
        self.message = message
        self.details = details or {}
        super().__init__(message)


class DataExporter:
    """Handles exporting of data to various file formats."""

    def __init__(self, logger: logging.Logger) -> None:
        """
        Initialize the exporter.

        Args:
            logger: Logger instance
        """
        self._logger = logger

    async def export_csv(
            self,
            data: List[Dict[str, Any]],
            columns: List[str],
            column_map: Dict[str, str],
            file_path: str
    ) -> None:
        """
        Export data to a CSV file.

        Args:
            data: List of data dictionaries
            columns: List of column IDs to include
            column_map: Mapping from column IDs to display names
            file_path: Path to the output file

        Raises:
            ExportError: If the export fails
        """
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=columns, extrasaction='ignore')

                # Write header row with display names
                writer.writerow({col: column_map.get(col, col) for col in columns})

                # Write data rows
                for row in data:
                    clean_row = {}
                    for col in columns:
                        value = row.get(col, '')
                        clean_row[col] = str(value) if value is not None else ''
                    writer.writerow(clean_row)

            self._logger.info(f'Exported {len(data)} rows to CSV: {file_path}')

        except Exception as e:
            self._logger.error(f'Failed to export CSV: {str(e)}')
            raise ExportError(f'Failed to export CSV: {str(e)}') from e

    async def export_excel(
            self,
            data: List[Dict[str, Any]],
            columns: List[str],
            column_map: Dict[str, str],
            file_path: str,
            sheet_name: Optional[str] = None
    ) -> None:
        """
        Export data to an Excel file.

        Args:
            data: List of data dictionaries
            columns: List of column IDs to include
            column_map: Mapping from column IDs to display names
            file_path: Path to the output file
            sheet_name: Optional name for the worksheet

        Raises:
            ExportError: If the export fails or Excel support is not available
        """
        if not EXCEL_AVAILABLE:
            raise ExportError('Excel export is not available. Please install openpyxl.')

        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Set sheet name
            if not sheet_name:
                sheet_name = 'VCdb Results'
            ws.title = sheet_name

            # Add timestamp
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=1, column=1, value=f'Generated: {timestamp}')
            ws.cell(row=1, column=1).font = Font(italic=True)

            # Header styles
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            header_alignment = Alignment(horizontal='center')

            # Write header row
            for col_idx, col_id in enumerate(columns, 1):
                cell = ws.cell(row=2, column=col_idx, value=column_map.get(col_id, col_id))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write data rows
            for row_idx, row_data in enumerate(data, 3):
                for col_idx, col_id in enumerate(columns, 1):
                    value = row_data.get(col_id, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col_idx, _ in enumerate(columns, 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                max_length = 0

                # Sample rows to determine width
                for row_idx in range(2, min(len(data) + 3, 100)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = max(max_length + 2, 10)
                ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

            # Freeze panes
            ws.freeze_panes = 'A3'

            # Save workbook
            wb.save(file_path)

            self._logger.info(f'Exported {len(data)} rows to Excel: {file_path}')

        except Exception as e:
            self._logger.error(f'Failed to export Excel: {str(e)}')
            raise ExportError(f'Failed to export Excel: {str(e)}') from e

    async def export_all_data(
            self,
            database_callback: Callable[..., Awaitable[Tuple[List[Dict[str, Any]], int]]],
            filter_panels: List[Dict[str, List[int]]],
            columns: List[str],
            column_map: Dict[str, str],
            file_path: str,
            format_type: str,
            max_rows: int = 0,
            table_filters: Optional[Dict[str, Any]] = None,
            sort_by: Optional[str] = None,
            sort_desc: bool = False,
            progress_callback: Optional[Callable[[int, int], Awaitable[Optional[bool]]]] = None
    ) -> int:
        """
        Export all matching data to a file.

        Args:
            database_callback: Function to call for retrieving data
            filter_panels: List of filter dictionaries from multiple panels
            columns: List of column IDs to include
            column_map: Mapping from column IDs to display names
            file_path: Path to the output file
            format_type: Export format (e.g., 'csv', 'excel')
            max_rows: Maximum number of rows to export (0 for all)
            table_filters: Additional table filters
            sort_by: Column to sort by
            sort_desc: Whether to sort in descending order
            progress_callback: Optional callback for reporting progress

        Returns:
            Number of rows exported

        Raises:
            ExportError: If the export fails
        """
        try:
            # Get total count
            results, total_count = await database_callback(
                filter_panels=filter_panels,
                columns=columns,
                page=1,
                page_size=1,
                sort_by=sort_by,
                sort_desc=sort_desc,
                table_filters=table_filters
            )

            # Determine how many rows to export
            if max_rows == 0:
                max_rows = total_count

            total_to_export = min(total_count, max_rows)

            if total_to_export == 0:
                self._logger.warning('No data to export')
                return 0

            # Export based on format
            if format_type == 'csv':
                return await self._export_all_csv(
                    database_callback,
                    filter_panels,
                    columns,
                    column_map,
                    file_path,
                    total_to_export,
                    table_filters,
                    sort_by,
                    sort_desc,
                    progress_callback
                )
            elif format_type == 'excel' and EXCEL_AVAILABLE:
                return await self._export_all_excel(
                    database_callback,
                    filter_panels,
                    columns,
                    column_map,
                    file_path,
                    total_to_export,
                    table_filters,
                    sort_by,
                    sort_desc,
                    progress_callback
                )
            else:
                raise ExportError(f'Unsupported export format: {format_type}')

        except Exception as e:
            self._logger.error(f'Failed to export data: {str(e)}')
            raise ExportError(f'Failed to export data: {str(e)}') from e

    async def _export_all_csv(
            self,
            database_callback: Callable[..., Awaitable[Tuple[List[Dict[str, Any]], int]]],
            filter_panels: List[Dict[str, List[int]]],
            columns: List[str],
            column_map: Dict[str, str],
            file_path: str,
            total_to_export: int,
            table_filters: Optional[Dict[str, Any]],
            sort_by: Optional[str],
            sort_desc: bool,
            progress_callback: Optional[Callable[[int, int], Awaitable[Optional[bool]]]]
    ) -> int:
        """
        Export all matching data to a CSV file.

        Args:
            database_callback: Function to call for retrieving data
            filter_panels: List of filter dictionaries from multiple panels
            columns: List of column IDs to include
            column_map: Mapping from column IDs to display names
            file_path: Path to the output file
            total_to_export: Total number of rows to export
            table_filters: Additional table filters
            sort_by: Column to sort by
            sort_desc: Whether to sort in descending order
            progress_callback: Optional callback for reporting progress

        Returns:
            Number of rows exported
        """
        with open(file_path, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=columns, extrasaction='ignore')

            # Write header row
            writer.writerow({col: column_map.get(col, col) for col in columns})

            # Write data rows in batches
            batch_size = 1000
            rows_exported = 0

            for page in range(1, (total_to_export + batch_size - 1) // batch_size + 1):
                # Report progress at batch start
                if progress_callback:
                    cancel = await progress_callback(-1, total_to_export)
                    if cancel is False:
                        return rows_exported

                # Fetch batch of data
                results, _ = await database_callback(
                    filter_panels=filter_panels,
                    columns=columns,
                    page=page,
                    page_size=batch_size,
                    sort_by=sort_by,
                    sort_desc=sort_desc,
                    table_filters=table_filters
                )

                # Process each row in batch
                for row in results:
                    clean_row = {}
                    for col in columns:
                        value = row.get(col, '')
                        clean_row[col] = str(value) if value is not None else ''
                    writer.writerow(clean_row)

                    rows_exported += 1

                    # Report progress for each row
                    if progress_callback:
                        cancel = await progress_callback(rows_exported, total_to_export)
                        if cancel is False:
                            return rows_exported

                    # Stop if we've exported enough rows
                    if rows_exported >= total_to_export:
                        break

        self._logger.info(f'Exported {rows_exported} rows to CSV: {file_path}')
        return rows_exported

    async def _export_all_excel(
            self,
            database_callback: Callable[..., Awaitable[Tuple[List[Dict[str, Any]], int]]],
            filter_panels: List[Dict[str, List[int]]],
            columns: List[str],
            column_map: Dict[str, str],
            file_path: str,
            total_to_export: int,
            table_filters: Optional[Dict[str, Any]],
            sort_by: Optional[str],
            sort_desc: bool,
            progress_callback: Optional[Callable[[int, int], Awaitable[Optional[bool]]]]
    ) -> int:
        """
        Export all matching data to an Excel file.

        Args:
            database_callback: Function to call for retrieving data
            filter_panels: List of filter dictionaries from multiple panels
            columns: List of column IDs to include
            column_map: Mapping from column IDs to display names
            file_path: Path to the output file
            total_to_export: Total number of rows to export
            table_filters: Additional table filters
            sort_by: Column to sort by
            sort_desc: Whether to sort in descending order
            progress_callback: Optional callback for reporting progress

        Returns:
            Number of rows exported
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'VCdb Results'

        # Add timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=1, column=1, value=f'Generated: {timestamp}')
        ws.cell(row=1, column=1).font = Font(italic=True)

        # Header styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        header_alignment = Alignment(horizontal='center')

        # Write header row
        for col_idx, col_id in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=column_map.get(col_id, col_id))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows in batches
        batch_size = 1000
        rows_exported = 0
        excel_row = 3  # Start after header

        for page in range(1, (total_to_export + batch_size - 1) // batch_size + 1):
            # Report progress at batch start
            if progress_callback:
                cancel = await progress_callback(-1, total_to_export)
                if cancel is False:
                    return rows_exported

            # Fetch batch of data
            results, _ = await database_callback(
                filter_panels=filter_panels,
                columns=columns,
                page=page,
                page_size=batch_size,
                sort_by=sort_by,
                sort_desc=sort_desc,
                table_filters=table_filters
            )

            # Process each row in batch
            for row_data in results:
                # Write row to Excel
                for col_idx, col_id in enumerate(columns, 1):
                    value = row_data.get(col_id, '')
                    ws.cell(row=excel_row, column=col_idx, value=value)

                excel_row += 1
                rows_exported += 1

                # Report progress for each row
                if progress_callback:
                    cancel = await progress_callback(rows_exported, total_to_export)
                    if cancel is False:
                        return rows_exported

                # Stop if we've exported enough rows
                if rows_exported >= total_to_export:
                    break

        # Auto-adjust column widths
        for col_idx, _ in enumerate(columns, 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0

            # Sample rows to determine width
            for row_idx in range(2, min(excel_row, 102)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max(max_length + 2, 10)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

        # Freeze panes
        ws.freeze_panes = 'A3'

        # Save workbook
        wb.save(file_path)

        self._logger.info(f'Exported {rows_exported} rows to Excel: {file_path}')
        return rows_exported